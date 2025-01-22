import openpyxl
import pandas as pd
import requests as req
import datetime

# app = Flask(__name__)
# CORS(app)

def read_period_from_excel(file_path):
    workbook = openpyxl.load_workbook('Book8.xlsx')
    sheet = workbook.active
    first_column_values=[]
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            first_column_values.append(cell.value)
    return first_column_values
def read_Activity_from_excel(file_path):
    workbook = openpyxl.load_workbook('Book8.xlsx')
    sheet = workbook.active
    Activity_column_values=[]
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
        for cell in row:
            Activity_column_values.append(cell.value)
    return Activity_column_values
def get_quater_dates(values):
    year, quater = values.split()
    year=int(year)
    if(quater=='Q1'):
        start_date=f"{year}-01-01"
        end_date=f"{year}-03-31"
    elif (quater=='Q2'):
        start_date = f"{year}-04-01"
        end_date = f"{year}-06-30"
    elif(quater=='Q3'):
        start_date = f"{year}-07-01"
        end_date = f"{year}-09-30"
    else:
        start_date = f"{year}-10-01"
        end_date = f"{year}-12-31"
    return start_date,end_date

def fetch_stock_data(symbol,values):
    # base_url = 'http://api.marketstack.com/v1/eod'
    # access_key = 'd17e9b08270b218b1b39304c52b43234'
    start_date, end_date = get_quater_dates(values)
    querystring = {"symbols":symbol,"date_from":start_date, "date_to":end_date}
    response = req.get('http://api.marketstack.com/v1/eod?access_key=d17e9b08270b218b1b39304c52b43234',params=querystring)
    if response.status_code ==200:
        data=response.json()
        return data['data'] if 'data' in data else[]
    else:
        print('error occured')
        return []
Period= read_period_from_excel(r'C:\Users\suman\PycharmProjects\pythonProject\.venv\Book8.xlsx')
workbook = openpyxl.load_workbook(r'C:\Users\suman\PycharmProjects\pythonProject\.venv\Book8.xlsx')
sheet = workbook.active
price_column = 'F'
row = 2
symbol = "AME"
Activity= read_Activity_from_excel(r'C:\Users\suman\PycharmProjects\pythonProject\.venv\Book8.xlsx')
for row, (values, values1) in enumerate(zip(Period, Activity), start=2):
    print(f"Processing row {row}: Period={values}, Activity={values1}")
    if values1 is not None:
        stock_data = fetch_stock_data(symbol, values)
        if stock_data:
            if values1.startswith(("Add","Buy")):
                prices = [item['low'] for item in stock_data if item['low'] is not None]
                if prices:
                    min_price = min(prices)
                    print(f"The minimum price is: ${min_price:.2f}")
                    sheet[f'{price_column}{row}'] = round(min_price, 2)
                else:
                    sheet[f'{price_column}{row}'] = 'N/A'
            elif values1.startswith("Reduce"):
                prices = [item['high'] for item in stock_data if item['high'] is not None]
                if prices:
                    max_price = max(prices)
                    print(f"The maximum price is: ${max_price:.2f}")
                    sheet[f'{price_column}{row}'] = round(max_price, 2)
                else:
                    sheet[f'{price_column}{row}'] = 'N/A'
                    print("error occured")
            #     sheet[f'{price_column}{row}'] = 'N/A'
            else:
             sheet[f'{price_column}{row}'] = 'N/A'
    row += 1
workbook.save('Book8.xlsx')
workbook = openpyxl.load_workbook(r'C:\Users\suman\PycharmProjects\pythonProject\.venv\Book8.xlsx')
sheet = workbook.active
cell_value=sheet['B2'].value
#print(cell_value)
column_to_subtract = 'B'
column_to_multiply = 'F'
start_row=2
total_sum=0
for row in range(start_row, sheet.max_row + 1):
    current_cell = f"{column_to_subtract}{row}"
    next_cell = f"{column_to_subtract}{row+1}"
    multiply_cell = f"{column_to_multiply}{row}"
    # result_cell = f"{result_column}{row}"

    current_value = sheet[current_cell].value
    next_value = sheet[next_cell].value
    multiply_value = sheet[multiply_cell].value
    if next_value is None:
        next_value=0
        subtraction_result =current_value-next_value
        final_result = subtraction_result * multiply_value
        total_sum+=final_result

    else:
        subtraction_result = current_value - next_value
        final_result = subtraction_result * multiply_value
        total_sum += final_result
cell_value=sheet['B2'].value
avg_cost_copy=total_sum/cell_value
column_letter = 'F'
new_row = sheet.max_row + 1
sheet[f'{column_letter}{sheet.max_row}'] = 'avg_cost_copy'
sheet[f'{column_letter}{new_row}'] = avg_cost_copy
workbook.save('Book8.xlsx')
# try:
#     workbook.save('sales.xlsx')
#     print("Workbook saved successfully")
# except Exception as e:
#     print(f"Error saving workbook: {e}")
#
#
#         elif response.status_code == 422:
#         print("Error 422: Unprocessable Entity. The request was well-formed but was unable to be followed due to semantic errors.")
#         print("Response content:", response.text)
# else:
#     print(f"Request failed with status code : {response.status_code}")
#     print("Response content:", response.text)